"""
HAL9000 — Knowledge System
Loads boot-time knowledge files, manages user uploads, and provides
chunked retrieval via BM25 keyword search.

Boot-time:  knowledge/*.txt and remote URLs (unchanged from v1)
Uploads:    knowledge/uploads/always/  — small files, always in prompt
            knowledge/uploads/indexed/ — chunked large files, searched on-demand

File extractors handle: txt, md, py, js, json, yaml, csv, pdf, docx, xlsx, images.
"""

import json
import math
import os
import re
import threading
import time
import urllib.error
import urllib.request
import uuid

from config import cfg

# ── Paths ─────────────────────────────────────────────────

_BASE_DIR = os.path.dirname(os.path.dirname(__file__))
KNOWLEDGE_DIR = os.path.join(_BASE_DIR, "knowledge")
SOURCES_FILE = os.path.join(KNOWLEDGE_DIR, "sources.txt")
UPLOADS_DIR = os.path.join(KNOWLEDGE_DIR, "uploads")
ALWAYS_DIR = os.path.join(UPLOADS_DIR, "always")
INDEXED_DIR = os.path.join(UPLOADS_DIR, "indexed")
MANIFEST_FILE = os.path.join(UPLOADS_DIR, "manifest.json")

SUPPORTED_EXTENSIONS = {".txt", ".md", ".markdown"}
FETCH_TIMEOUT = 10

# Configurable thresholds (chars)
ALWAYS_MAX_KB = getattr(cfg, "KNOWLEDGE_ALWAYS_MAX_KB", 2)
CHUNK_SIZE = getattr(cfg, "KNOWLEDGE_CHUNK_SIZE", 1500)
CHUNK_OVERLAP = 200
MAX_CHUNKS_IN_PROMPT = getattr(cfg, "KNOWLEDGE_MAX_CHUNKS_IN_PROMPT", 3)
MAX_TOTAL_MB = getattr(cfg, "KNOWLEDGE_MAX_TOTAL_MB", 50)

# Supported upload file types
UPLOAD_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".csv", ".json", ".yaml", ".yml",
    ".py", ".js", ".ts", ".jsx", ".tsx", ".java", ".go", ".rs",
    ".c", ".cpp", ".h", ".hpp", ".rb", ".php", ".swift", ".kt",
    ".html", ".css", ".xml", ".sql", ".sh", ".bash", ".zsh",
    ".pdf", ".docx", ".xlsx",
    ".png", ".jpg", ".jpeg", ".gif", ".webp",
}


# ── Boot-time Loaders (v1, unchanged) ────────────────────

def load_local_files() -> list[tuple[str, str]]:
    """Read all .md and .txt files from knowledge/ (excluding sources.txt)."""
    entries = []
    if not os.path.isdir(KNOWLEDGE_DIR):
        return entries

    for filename in sorted(os.listdir(KNOWLEDGE_DIR)):
        if filename == "sources.txt":
            continue

        _, ext = os.path.splitext(filename)
        if ext.lower() not in SUPPORTED_EXTENSIONS:
            continue

        filepath = os.path.join(KNOWLEDGE_DIR, filename)
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                content = f.read().strip()
            if content:
                entries.append((filename, content))
                print(f"[Knowledge] Loaded: {filename} ({len(content)} chars)")
        except Exception as e:
            print(f"[Knowledge] Failed to read {filename}: {e}")

    return entries


def fetch_remote_sources() -> list[tuple[str, str]]:
    """Fetch llms.txt (or any text) from URLs listed in sources.txt."""
    entries = []
    if not os.path.isfile(SOURCES_FILE):
        return entries

    with open(SOURCES_FILE, "r", encoding="utf-8") as f:
        lines = f.readlines()

    for line in lines:
        url = line.strip()
        if not url or url.startswith("#"):
            continue

        try:
            req = urllib.request.Request(
                url,
                headers={"User-Agent": "HAL9000-KnowledgeLoader/1.0"},
            )
            with urllib.request.urlopen(req, timeout=FETCH_TIMEOUT) as resp:
                content = resp.read().decode("utf-8", errors="replace").strip()

            if content:
                label = url.split("//", 1)[-1]
                entries.append((label, content))
                print(f"[Knowledge] Fetched: {label} ({len(content)} chars)")
        except (urllib.error.URLError, Exception) as e:
            print(f"[Knowledge] Failed to fetch {url}: {e}")

    return entries


def load_all() -> str:
    """Load all knowledge and return as a single formatted string."""
    local = load_local_files()
    remote = fetch_remote_sources()
    all_entries = local + remote

    if not all_entries:
        print("[Knowledge] No knowledge files found.")
        return ""

    sections = []
    for label, content in all_entries:
        sections.append(f"--- {label} ---\n{content}")

    combined = "\n\n".join(sections)
    print(f"[Knowledge] Total: {len(all_entries)} sources, {len(combined)} chars")
    return combined


# ── File Extractors ──────────────────────────────────────

def _extract_text(filepath: str) -> str:
    """Read plain text files."""
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


def _extract_pdf(filepath: str) -> str:
    """Extract text from PDF using PyMuPDF."""
    try:
        import fitz  # pymupdf
    except ImportError:
        return "[PDF extraction requires pymupdf. Install: pip install pymupdf]"

    text_parts = []
    with fitz.open(filepath) as doc:
        for page in doc:
            text_parts.append(page.get_text())
    return "\n".join(text_parts)


def _extract_docx(filepath: str) -> str:
    """Extract text from Word document."""
    try:
        import docx
    except ImportError:
        return "[DOCX extraction requires python-docx. Install: pip install python-docx]"

    doc = docx.Document(filepath)
    return "\n".join(p.text for p in doc.paragraphs)


def _extract_xlsx(filepath: str) -> str:
    """Extract data from Excel spreadsheet."""
    try:
        import openpyxl
    except ImportError:
        return "[XLSX extraction requires openpyxl. Install: pip install openpyxl]"

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            lines.append("\t".join(cells))
    wb.close()
    return "\n".join(lines)


def _extract_csv(filepath: str) -> str:
    """Read CSV file as text."""
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        return f.read()


def _extract_image(filepath: str) -> str:
    """For images, return a placeholder — actual OCR would need vision LLM."""
    return f"[Image file: {os.path.basename(filepath)} — use vision to analyze]"


def extract_content(filepath: str) -> str:
    """Extract text content from a file based on its extension."""
    _, ext = os.path.splitext(filepath)
    ext = ext.lower()

    if ext == ".pdf":
        return _extract_pdf(filepath)
    elif ext == ".docx":
        return _extract_docx(filepath)
    elif ext == ".xlsx":
        return _extract_xlsx(filepath)
    elif ext == ".csv":
        return _extract_csv(filepath)
    elif ext in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
        return _extract_image(filepath)
    else:
        # All text-based files (code, markdown, json, yaml, etc.)
        return _extract_text(filepath)


# ── Chunking ─────────────────────────────────────────────

def _split_into_chunks(text: str, chunk_size: int = CHUNK_SIZE,
                       overlap: int = CHUNK_OVERLAP) -> list[str]:
    """Split text into overlapping chunks, preferring paragraph/sentence breaks."""
    if len(text) <= chunk_size:
        return [text]

    chunks = []
    start = 0

    while start < len(text):
        end = start + chunk_size

        if end >= len(text):
            chunks.append(text[start:])
            break

        # Try to break at paragraph
        para_break = text.rfind("\n\n", start + chunk_size // 2, end)
        if para_break > start:
            end = para_break + 2
        else:
            # Try sentence break
            sentence_break = max(
                text.rfind(". ", start + chunk_size // 2, end),
                text.rfind(".\n", start + chunk_size // 2, end),
                text.rfind("? ", start + chunk_size // 2, end),
                text.rfind("! ", start + chunk_size // 2, end),
            )
            if sentence_break > start:
                end = sentence_break + 2

        chunks.append(text[start:end])
        start = end - overlap  # overlap for continuity

    return chunks


# ── BM25 Search ──────────────────────────────────────────

def _tokenize(text: str) -> list[str]:
    """Simple word tokenizer — lowercase, strip punctuation."""
    return re.findall(r'[a-z0-9]+', text.lower())


def _build_bm25_index(chunks: list[str]) -> dict:
    """Build a BM25-ready index for a set of chunks."""
    doc_count = len(chunks)
    doc_lengths = []
    tf_per_doc = []  # term frequencies per document
    df = {}  # document frequency per term

    for chunk in chunks:
        tokens = _tokenize(chunk)
        doc_lengths.append(len(tokens))

        tf = {}
        seen = set()
        for token in tokens:
            tf[token] = tf.get(token, 0) + 1
            if token not in seen:
                df[token] = df.get(token, 0) + 1
                seen.add(token)
        tf_per_doc.append(tf)

    avg_dl = sum(doc_lengths) / max(doc_count, 1)

    return {
        "doc_count": doc_count,
        "doc_lengths": doc_lengths,
        "tf": tf_per_doc,
        "df": df,
        "avg_dl": avg_dl,
    }


def _bm25_score(query_tokens: list[str], doc_idx: int, index: dict,
                k1: float = 1.5, b: float = 0.75) -> float:
    """Calculate BM25 score for a single document."""
    score = 0.0
    n = index["doc_count"]
    dl = index["doc_lengths"][doc_idx]
    avg_dl = index["avg_dl"]
    tf_doc = index["tf"][doc_idx]
    df = index["df"]

    for term in query_tokens:
        if term not in tf_doc:
            continue
        tf_val = tf_doc[term]
        df_val = df.get(term, 0)
        # IDF component
        idf = math.log((n - df_val + 0.5) / (df_val + 0.5) + 1)
        # TF component with length normalization
        tf_norm = (tf_val * (k1 + 1)) / (tf_val + k1 * (1 - b + b * dl / avg_dl))
        score += idf * tf_norm

    return score


def search_chunks(query: str, chunks: list[str], index: dict,
                  top_k: int = MAX_CHUNKS_IN_PROMPT) -> list[tuple[int, float, str]]:
    """Search chunks using BM25, return top-k results as (index, score, text)."""
    query_tokens = _tokenize(query)
    if not query_tokens:
        return []

    scores = []
    for i in range(len(chunks)):
        s = _bm25_score(query_tokens, i, index)
        if s > 0:
            scores.append((i, s, chunks[i]))

    scores.sort(key=lambda x: -x[1])
    return scores[:top_k]


# ── Manifest (Upload Index) ─────────────────────────────

_manifest_lock = threading.Lock()


def _ensure_dirs():
    """Create upload directories if they don't exist."""
    os.makedirs(ALWAYS_DIR, exist_ok=True)
    os.makedirs(INDEXED_DIR, exist_ok=True)


def _load_manifest() -> list[dict]:
    """Load the upload manifest."""
    if not os.path.isfile(MANIFEST_FILE):
        return []
    try:
        with open(MANIFEST_FILE, "r") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return []


def _save_manifest(entries: list[dict]):
    """Save the upload manifest."""
    _ensure_dirs()
    with open(MANIFEST_FILE, "w") as f:
        json.dump(entries, f, indent=2)


def _total_upload_size_mb() -> float:
    """Calculate total size of uploads directory in MB."""
    total = 0
    for root, _, files in os.walk(UPLOADS_DIR):
        for fname in files:
            total += os.path.getsize(os.path.join(root, fname))
    return total / (1024 * 1024)


# ── Upload API ───────────────────────────────────────────

def upload_file(filepath: str, original_name: str,
                mode: str = "auto") -> dict:
    """
    Process and store an uploaded file.

    Args:
        filepath: Path to the uploaded temp file
        original_name: Original filename
        mode: "auto" (decide by size), "deep" (LLM summarize chunks),
              "skim" (keyword index only), "always" (force always-loaded)

    Returns:
        dict with upload result info
    """
    _ensure_dirs()

    _, ext = os.path.splitext(original_name)
    ext = ext.lower()
    if ext not in UPLOAD_EXTENSIONS:
        return {"error": f"Unsupported file type: {ext}"}

    # Check storage limit
    if _total_upload_size_mb() >= MAX_TOTAL_MB:
        return {"error": f"Storage limit reached ({MAX_TOTAL_MB} MB). Delete some files first."}

    # Extract content
    content = extract_content(filepath)
    if not content or content.startswith("["):
        # Image or extraction failed — store reference only
        if ext in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
            pass  # images stored as-is for vision
        elif content.startswith("["):
            return {"error": content}

    content_size_kb = len(content.encode("utf-8")) / 1024
    file_id = str(uuid.uuid4())[:8]

    with _manifest_lock:
        manifest = _load_manifest()

        # Determine storage mode
        if mode == "auto":
            if content_size_kb < ALWAYS_MAX_KB:
                mode = "always"
            else:
                # Return "ask" signal — caller should prompt user
                return {
                    "id": file_id,
                    "name": original_name,
                    "size_kb": round(content_size_kb, 1),
                    "needs_choice": True,
                    "content": content,  # pass back for re-processing
                }
        elif mode == "always":
            pass  # force always-loaded

        if mode == "always":
            # Save to always/ directory
            dest = os.path.join(ALWAYS_DIR, f"{file_id}_{original_name}")
            with open(dest, "w", encoding="utf-8") as f:
                f.write(content)

            entry = {
                "id": file_id,
                "name": original_name,
                "mode": "always",
                "size_kb": round(content_size_kb, 1),
                "uploaded_at": time.time(),
                "path": dest,
            }
            manifest.append(entry)
            _save_manifest(manifest)

            print(f"[Knowledge] Uploaded (always): {original_name} ({content_size_kb:.1f} KB)")
            return {"id": file_id, "name": original_name, "mode": "always",
                    "size_kb": round(content_size_kb, 1)}

        # Chunked storage (deep or skim)
        chunks = _split_into_chunks(content)
        chunk_dir = os.path.join(INDEXED_DIR, file_id)
        os.makedirs(chunk_dir, exist_ok=True)

        # Save chunks
        for i, chunk in enumerate(chunks):
            chunk_path = os.path.join(chunk_dir, f"chunk_{i:03d}.txt")
            with open(chunk_path, "w", encoding="utf-8") as f:
                f.write(chunk)

        # Build and save BM25 index
        bm25_index = _build_bm25_index(chunks)
        index_path = os.path.join(chunk_dir, "bm25_index.json")
        with open(index_path, "w") as f:
            json.dump(bm25_index, f)

        # Save metadata
        meta = {
            "id": file_id,
            "name": original_name,
            "mode": mode,
            "chunks": len(chunks),
            "size_kb": round(content_size_kb, 1),
            "uploaded_at": time.time(),
        }
        meta_path = os.path.join(chunk_dir, "meta.json")
        with open(meta_path, "w") as f:
            json.dump(meta, f, indent=2)

        entry = {
            "id": file_id,
            "name": original_name,
            "mode": mode,
            "size_kb": round(content_size_kb, 1),
            "chunks": len(chunks),
            "uploaded_at": time.time(),
            "path": chunk_dir,
        }
        manifest.append(entry)
        _save_manifest(manifest)

        print(f"[Knowledge] Uploaded ({mode}): {original_name} "
              f"({content_size_kb:.1f} KB, {len(chunks)} chunks)")
        return {"id": file_id, "name": original_name, "mode": mode,
                "chunks": len(chunks), "size_kb": round(content_size_kb, 1)}


def process_upload_with_mode(file_id: str, content: str, original_name: str,
                             mode: str) -> dict:
    """Process a file that was deferred pending user mode choice (deep/skim)."""
    _ensure_dirs()

    content_size_kb = len(content.encode("utf-8")) / 1024

    chunks = _split_into_chunks(content)
    chunk_dir = os.path.join(INDEXED_DIR, file_id)
    os.makedirs(chunk_dir, exist_ok=True)

    for i, chunk in enumerate(chunks):
        chunk_path = os.path.join(chunk_dir, f"chunk_{i:03d}.txt")
        with open(chunk_path, "w", encoding="utf-8") as f:
            f.write(chunk)

    bm25_index = _build_bm25_index(chunks)
    index_path = os.path.join(chunk_dir, "bm25_index.json")
    with open(index_path, "w") as f:
        json.dump(bm25_index, f)

    meta = {
        "id": file_id,
        "name": original_name,
        "mode": mode,
        "chunks": len(chunks),
        "size_kb": round(content_size_kb, 1),
        "uploaded_at": time.time(),
    }
    meta_path = os.path.join(chunk_dir, "meta.json")
    with open(meta_path, "w") as f:
        json.dump(meta, f, indent=2)

    with _manifest_lock:
        manifest = _load_manifest()
        manifest.append({
            "id": file_id,
            "name": original_name,
            "mode": mode,
            "size_kb": round(content_size_kb, 1),
            "chunks": len(chunks),
            "uploaded_at": time.time(),
            "path": chunk_dir,
        })
        _save_manifest(manifest)

    print(f"[Knowledge] Uploaded ({mode}): {original_name} "
          f"({content_size_kb:.1f} KB, {len(chunks)} chunks)")
    return {"id": file_id, "name": original_name, "mode": mode,
            "chunks": len(chunks), "size_kb": round(content_size_kb, 1)}


# ── Retrieval ────────────────────────────────────────────

def load_always_files() -> str:
    """Load all always-loaded upload files as a combined string."""
    if not os.path.isdir(ALWAYS_DIR):
        return ""

    sections = []
    for filename in sorted(os.listdir(ALWAYS_DIR)):
        filepath = os.path.join(ALWAYS_DIR, filename)
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                content = f.read().strip()
            if content:
                # Strip the UUID prefix for display
                display_name = "_".join(filename.split("_")[1:]) if "_" in filename else filename
                sections.append(f"--- {display_name} ---\n{content}")
        except Exception:
            pass

    return "\n\n".join(sections)


def recall(query: str, top_k: int = MAX_CHUNKS_IN_PROMPT) -> list[dict]:
    """Search all indexed knowledge files for relevant chunks."""
    if not os.path.isdir(INDEXED_DIR):
        return []

    all_results = []

    for file_id in os.listdir(INDEXED_DIR):
        chunk_dir = os.path.join(INDEXED_DIR, file_id)
        if not os.path.isdir(chunk_dir):
            continue

        # Load metadata
        meta_path = os.path.join(chunk_dir, "meta.json")
        if not os.path.isfile(meta_path):
            continue
        try:
            with open(meta_path, "r") as f:
                meta = json.load(f)
        except (json.JSONDecodeError, IOError):
            continue

        # Load chunks
        chunk_files = sorted(
            f for f in os.listdir(chunk_dir)
            if f.startswith("chunk_") and f.endswith(".txt")
        )
        chunks = []
        for cf in chunk_files:
            with open(os.path.join(chunk_dir, cf), "r", encoding="utf-8") as f:
                chunks.append(f.read())

        if not chunks:
            continue

        # Load or rebuild BM25 index
        index_path = os.path.join(chunk_dir, "bm25_index.json")
        if os.path.isfile(index_path):
            try:
                with open(index_path, "r") as f:
                    bm25_index = json.load(f)
            except (json.JSONDecodeError, IOError):
                bm25_index = _build_bm25_index(chunks)
        else:
            bm25_index = _build_bm25_index(chunks)

        # Search
        results = search_chunks(query, chunks, bm25_index, top_k=top_k)
        for idx, score, text in results:
            all_results.append({
                "file_id": file_id,
                "file_name": meta.get("name", file_id),
                "chunk_index": idx,
                "score": round(score, 3),
                "text": text,
            })

    # Sort by score across all files, return top_k
    all_results.sort(key=lambda x: -x["score"])
    return all_results[:top_k]


def list_uploads() -> list[dict]:
    """List all uploaded knowledge files."""
    manifest = _load_manifest()
    result = []
    for entry in manifest:
        result.append({
            "id": entry["id"],
            "name": entry["name"],
            "mode": entry["mode"],
            "size_kb": entry.get("size_kb", 0),
            "chunks": entry.get("chunks", 0),
            "uploaded_at": entry.get("uploaded_at", 0),
        })
    return result


def delete_upload(file_id: str) -> bool:
    """Delete an uploaded knowledge file and its chunks."""
    import shutil

    with _manifest_lock:
        manifest = _load_manifest()
        entry = None
        for e in manifest:
            if e["id"] == file_id:
                entry = e
                break

        if not entry:
            return False

        # Delete files
        path = entry.get("path", "")
        if path and os.path.exists(path):
            if os.path.isdir(path):
                shutil.rmtree(path)
            else:
                os.remove(path)

        # Remove from manifest
        manifest = [e for e in manifest if e["id"] != file_id]
        _save_manifest(manifest)

        print(f"[Knowledge] Deleted: {entry['name']} ({file_id})")
        return True


def get_storage_info() -> dict:
    """Get storage usage info."""
    total_mb = _total_upload_size_mb()
    count = len(_load_manifest())
    return {
        "used_mb": round(total_mb, 2),
        "max_mb": MAX_TOTAL_MB,
        "file_count": count,
    }
